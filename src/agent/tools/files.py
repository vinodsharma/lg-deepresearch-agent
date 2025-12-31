"""File analysis tools for PDFs and documents."""

import fitz  # PyMuPDF
from docx import Document
from langchain_core.tools import tool
from openpyxl import load_workbook


@tool
def analyze_pdf(file_path: str, max_pages: int = 50) -> str:
    """Extract text content from a PDF file.

    Args:
        file_path: Path to the PDF file.
        max_pages: Maximum number of pages to process (default 50).

    Returns:
        Extracted text content from the PDF.
    """
    try:
        doc = fitz.open(file_path)

        pages_to_process = min(len(doc), max_pages)
        text_parts = []

        for i, page in enumerate(doc):
            if i >= pages_to_process:
                break
            text = page.get_text()
            if text.strip():
                text_parts.append(f"--- Page {i + 1} ---\n{text}")

        doc_length = len(doc)
        doc.close()

        if not text_parts:
            return "No text content found in PDF."

        result = "\n\n".join(text_parts)

        if doc_length > max_pages:
            result += f"\n\n[Truncated: showing {max_pages} of {doc_length} pages]"

        return result

    except Exception as e:
        return f"Error analyzing PDF: {str(e)}"


@tool
def analyze_document(file_path: str) -> str:
    """Extract content from DOCX or XLSX files.

    Args:
        file_path: Path to the document file.

    Returns:
        Extracted content from the document.
    """
    try:
        if file_path.endswith(".docx"):
            return _analyze_docx(file_path)
        elif file_path.endswith(".xlsx"):
            return _analyze_xlsx(file_path)
        else:
            return f"Unsupported file type: {file_path}"
    except Exception as e:
        return f"Error analyzing document: {str(e)}"


def _analyze_docx(file_path: str) -> str:
    """Extract text from DOCX file."""
    doc = Document(file_path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def _analyze_xlsx(file_path: str) -> str:
    """Extract data from XLSX file."""
    wb = load_workbook(file_path, read_only=True)

    results = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        results.append(f"## Sheet: {sheet_name}\n")

        rows = []
        for row in sheet.iter_rows(values_only=True):
            row_str = " | ".join(str(cell) if cell else "" for cell in row)
            if row_str.strip(" |"):
                rows.append(row_str)

        results.append("\n".join(rows[:100]))  # Limit rows

        if len(rows) > 100:
            results.append(f"\n[Truncated: showing 100 of {len(rows)} rows]")

    wb.close()
    return "\n\n".join(results)
